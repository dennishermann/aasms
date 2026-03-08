"""Export service for generating full dataset exports in Excel format."""

from typing import Dict, List, Optional, Any
from io import BytesIO
from datetime import datetime
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import text

logger = logging.getLogger(__name__)


class ExportService:
    """Service for generating comprehensive Excel exports of SMS data."""

    def __init__(self, db: Session):
        """
        Initialize export service.

        Args:
            db: SQLAlchemy database session
        """
        self.db = db

    def generate_full_export(self, study_id: str) -> bytes:
        """
        Generate a comprehensive multi-sheet Excel export of the study.

        Args:
            study_id: The study ID to export

        Returns:
            Binary Excel file content (.xlsx)
        """
        # Fetch all study data
        study = self._fetch_study(study_id)
        if not study:
            raise ValueError(f"Study {study_id} not found")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_protocol_sheet(wb, study)
        self._create_all_sources_sheet(wb, study)
        self._create_included_sources_sheet(wb, study)
        self._create_exclusion_log_sheet(wb, study)
        self._create_classification_matrix_sheet(wb, study)
        self._create_data_dictionary_sheet(wb, study)

        # Write to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _fetch_study(self, study_id: str) -> Optional[Dict[str, Any]]:
        """Fetch study data with all relationships."""
        query = text("""
            SELECT
                s.id, s.title, s.description, s.motivation, s.status, s."createdAt",
                COALESCE(json_agg(
                    json_build_object(
                        'id', rq.id,
                        'question', rq.question,
                        'order', rq.order
                    ) ORDER BY rq."order"
                ) FILTER (WHERE rq.id IS NOT NULL), '[]'::json) as "researchQuestions",
                COALESCE(json_agg(
                    json_build_object(
                        'id', ic.id,
                        'criterion', ic.criterion,
                        'order', ic."order"
                    ) ORDER BY ic."order"
                ) FILTER (WHERE ic.id IS NOT NULL), '[]'::json) as "inclusionCriteria",
                COALESCE(json_agg(
                    json_build_object(
                        'id', ec.id,
                        'criterion', ec.criterion,
                        'order', ec."order"
                    ) ORDER BY ec."order"
                ) FILTER (WHERE ec.id IS NOT NULL), '[]'::json) as "exclusionCriteria"
            FROM "Study" s
            LEFT JOIN "ResearchQuestion" rq ON s.id = rq."studyId"
            LEFT JOIN "StudyParameters" sp ON s.id = sp."studyId"
            LEFT JOIN "InclusionCriterion" ic ON sp.id = ic."parametersId"
            LEFT JOIN "ExclusionCriterion" ec ON sp.id = ec."parametersId"
            WHERE s.id = :study_id
            GROUP BY s.id, s.title, s.description, s.motivation, s.status, s."createdAt"
        """)
        result = self.db.execute(query, {"study_id": study_id}).fetchone()
        if not result:
            return None
        return dict(result._mapping)

    def _fetch_facets(self, study_id: str) -> List[Dict[str, Any]]:
        """Fetch facets with categories for a study."""
        query = text("""
            SELECT
                f.id, f.name, f.description, f.type, f."order",
                COALESCE(json_agg(
                    json_build_object(
                        'id', fc.id,
                        'name', fc.name,
                        'description', fc.description,
                        'order', fc."order"
                    ) ORDER BY fc."order"
                ) FILTER (WHERE fc.id IS NOT NULL), '[]'::json) as categories
            FROM "Facet" f
            LEFT JOIN "FacetCategory" fc ON f.id = fc."facetId"
            WHERE f."studyId" = :study_id
            GROUP BY f.id, f.name, f.description, f.type, f."order"
            ORDER BY f."order", f.name
        """)
        results = self.db.execute(query, {"study_id": study_id}).fetchall()
        return [dict(r._mapping) for r in results]

    def _fetch_all_sources(self, study_id: str) -> List[Dict[str, Any]]:
        """Fetch all sources for a study."""
        query = text("""
            SELECT
                s.id, s.title, s.authors, s.year, s."publicationDate", s.venue,
                s."venueType", s.doi, s.url, s."sourceCategory", s."greyLiteratureTier",
                s.status, s."finalDecision", sa."inclusionRecommendation",
                sa."confidenceScore", sa."exclusionReasoning", sa."inclusionReasoning"
            FROM "Source" s
            LEFT JOIN "SourceAnalysis" sa ON s.id = sa."sourceId"
            WHERE s."studyId" = :study_id
            ORDER BY s.title
        """)
        results = self.db.execute(query, {"study_id": study_id}).fetchall()
        return [dict(r._mapping) for r in results]

    def _fetch_included_sources(self, study_id: str) -> List[Dict[str, Any]]:
        """Fetch included sources with their classifications."""
        query = text("""
            SELECT
                s.id, s.title, s.authors, s.year, s."publicationDate", s.venue, s.doi,
                COALESCE(json_object_agg(
                    f.id, json_build_object(
                        'facetName', f.name,
                        'facetType', f.type,
                        'classifications', COALESCE(json_agg(
                            json_build_object(
                                'categoryId', c."categoryId",
                                'categoryName', fc.name,
                                'value', c.value,
                                'confidence', c.confidence
                            )
                        ) FILTER (WHERE c.id IS NOT NULL), '[]'::json)
                    )
                ), '{}'::json) as classifications
            FROM "Source" s
            LEFT JOIN "SourceAnalysis" sa ON s.id = sa."sourceId"
            LEFT JOIN "Classification" c ON sa.id = c."analysisId"
            LEFT JOIN "Facet" f ON c."facetId" = f.id
            LEFT JOIN "FacetCategory" fc ON c."categoryId" = fc.id
            WHERE s."studyId" = :study_id
            AND (s."finalDecision" = 'INCLUDE' OR s.status = 'CLASSIFIED')
            GROUP BY s.id, s.title, s.authors, s.year, s."publicationDate", s.venue, s.doi
            ORDER BY s.title
        """)
        results = self.db.execute(query, {"study_id": study_id}).fetchall()
        return [dict(r._mapping) for r in results]

    def _fetch_excluded_sources(self, study_id: str) -> List[Dict[str, Any]]:
        """Fetch excluded sources."""
        query = text("""
            SELECT
                s.title, s.authors, s.year, s."publicationDate",
                sa."exclusionReasoning", sa."inclusionReasoning",
                sa."confidenceScore"
            FROM "Source" s
            LEFT JOIN "SourceAnalysis" sa ON s.id = sa."sourceId"
            WHERE s."studyId" = :study_id
            AND s."finalDecision" = 'EXCLUDE'
            ORDER BY s.title
        """)
        results = self.db.execute(query, {"study_id": study_id}).fetchall()
        return [dict(r._mapping) for r in results]

    def _create_protocol_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create Protocol sheet with study metadata."""
        ws = wb.create_sheet("Protocol", 0)

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70

        row = 1

        # Title
        self._add_row(ws, row, "Study Title", study.get("title", ""), bold=True)
        row += 1

        # Description
        self._add_row(ws, row, "Description", study.get("description", "") or "", bold=True)
        row += 1

        # Motivation
        self._add_row(ws, row, "Motivation", study.get("motivation", "") or "", bold=True)
        row += 1

        # Status
        self._add_row(ws, row, "Status", study.get("status", ""), bold=True)
        row += 2

        # Research Questions
        ws.cell(row, 1, "Research Questions").font = Font(bold=True, size=12)
        row += 1

        rqs = study.get("researchQuestions", [])
        for idx, rq in enumerate(rqs, 1):
            ws.cell(row, 1, f"{idx}.")
            ws.cell(row, 2, rq.get("question", ""))
            row += 1

        if not rqs:
            ws.cell(row, 2, "No research questions defined")
            row += 1

        row += 1

        # Inclusion Criteria
        ws.cell(row, 1, "Inclusion Criteria").font = Font(bold=True, size=12)
        row += 1

        inc_criteria = study.get("inclusionCriteria", [])
        for idx, ic in enumerate(inc_criteria, 1):
            ws.cell(row, 1, f"{idx}.")
            ws.cell(row, 2, ic.get("criterion", ""))
            row += 1

        if not inc_criteria:
            ws.cell(row, 2, "No inclusion criteria defined")
            row += 1

        row += 1

        # Exclusion Criteria
        ws.cell(row, 1, "Exclusion Criteria").font = Font(bold=True, size=12)
        row += 1

        exc_criteria = study.get("exclusionCriteria", [])
        for idx, ec in enumerate(exc_criteria, 1):
            ws.cell(row, 1, f"{idx}.")
            ws.cell(row, 2, ec.get("criterion", ""))
            row += 1

        if not exc_criteria:
            ws.cell(row, 2, "No exclusion criteria defined")
            row += 1

        row += 2

        # Facet Definitions
        ws.cell(row, 1, "Facet Definitions").font = Font(bold=True, size=12)
        row += 1

        facets = self._fetch_facets(study["id"])

        for facet in facets:
            ws.cell(row, 1, f"Facet: {facet.get('name', '')}").font = Font(bold=True)
            row += 1

            ws.cell(row, 1, "Type")
            ws.cell(row, 2, facet.get("type", ""))
            row += 1

            if facet.get("description"):
                ws.cell(row, 1, "Description")
                ws.cell(row, 2, facet.get("description", ""))
                row += 1

            categories = facet.get("categories", [])
            if categories and facet.get("type") in ["CLOSED", "OPEN_CODED"]:
                ws.cell(row, 1, "Categories").font = Font(italic=True)
                row += 1
                for cat in categories:
                    ws.cell(row, 1, f"  - {cat.get('name', '')}")
                    if cat.get("description"):
                        ws.cell(row, 2, cat.get("description", ""))
                    row += 1

            row += 1

        # Generated date
        row += 1
        ws.cell(row, 1, "Date Generated")
        ws.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    def _create_all_sources_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create All Sources sheet with all sources."""
        ws = wb.create_sheet("All Sources", 1)

        # Headers
        headers = [
            "ID", "Title", "Authors", "Year", "Venue", "Venue Type", "DOI", "URL",
            "Source Category", "Grey Lit Tier", "Status", "Final Decision",
            "Inclusion Confidence", "Exclusion Reasoning"
        ]

        self._add_header_row(ws, headers)

        # Data
        sources = self._fetch_all_sources(study["id"])

        for row_idx, source in enumerate(sources, 2):
            ws.cell(row_idx, 1, source.get("id", ""))
            ws.cell(row_idx, 2, source.get("title", ""))
            ws.cell(row_idx, 3, "; ".join(source.get("authors", []) or []))

            year = None
            pub_date = source.get("publicationDate")
            if pub_date:
                year = pub_date.year if hasattr(pub_date, 'year') else str(pub_date)[:4]
            ws.cell(row_idx, 4, year or "")

            ws.cell(row_idx, 5, source.get("venue", "") or "")
            ws.cell(row_idx, 6, source.get("venueType", "") or "")
            ws.cell(row_idx, 7, source.get("doi", "") or "")
            ws.cell(row_idx, 8, source.get("url", "") or "")
            ws.cell(row_idx, 9, source.get("sourceCategory", "") or "")
            ws.cell(row_idx, 10, source.get("greyLiteratureTier", "") or "")
            ws.cell(row_idx, 11, source.get("status", "") or "")
            ws.cell(row_idx, 12, source.get("finalDecision", "") or "")
            ws.cell(row_idx, 13, source.get("confidenceScore") or "")
            ws.cell(row_idx, 14, source.get("exclusionReasoning", "") or "")

        # Format columns
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_included_sources_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create Included Sources sheet with facet classifications."""
        ws = wb.create_sheet("Included Sources", 2)

        facets = self._fetch_facets(study["id"])
        sources = self._fetch_included_sources(study["id"])

        # Headers: ID, Title, Authors, Year, Venue, DOI, then facets
        headers = ["ID", "Title", "Authors", "Year", "Venue", "DOI"] + [f.get("name", "") for f in facets]

        self._add_header_row(ws, headers)

        # Data
        for row_idx, source in enumerate(sources, 2):
            ws.cell(row_idx, 1, source.get("id", ""))
            ws.cell(row_idx, 2, source.get("title", ""))
            ws.cell(row_idx, 3, "; ".join(source.get("authors", []) or []))

            year = None
            pub_date = source.get("publicationDate")
            if pub_date:
                year = pub_date.year if hasattr(pub_date, 'year') else str(pub_date)[:4]
            ws.cell(row_idx, 4, year or "")

            ws.cell(row_idx, 5, source.get("venue", "") or "")
            ws.cell(row_idx, 6, source.get("doi", "") or "")

            # Facet classifications
            classifications = source.get("classifications", {})
            for col_idx, facet in enumerate(facets, 7):
                facet_data = classifications.get(facet["id"], {})
                facet_type = facet.get("type", "CLOSED")

                if facet_data:
                    class_list = facet_data.get("classifications", [])
                    if facet_type == "OPEN":
                        # For OPEN facets, show the value
                        values = [c.get("value", "") for c in class_list if c.get("value")]
                        ws.cell(row_idx, col_idx, "; ".join(values))
                    else:
                        # For CLOSED/OPEN_CODED, show category names
                        cat_names = [c.get("categoryName", "") for c in class_list if c.get("categoryName")]
                        ws.cell(row_idx, col_idx, "; ".join(cat_names))

        # Format columns
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _create_exclusion_log_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create Exclusion Log sheet."""
        ws = wb.create_sheet("Exclusion Log", 3)

        headers = ["Title", "Authors", "Year", "Reason for Exclusion", "Confidence Score"]

        self._add_header_row(ws, headers)

        sources = self._fetch_excluded_sources(study["id"])

        for row_idx, source in enumerate(sources, 2):
            ws.cell(row_idx, 1, source.get("title", ""))
            ws.cell(row_idx, 2, "; ".join(source.get("authors", []) or []))

            year = None
            pub_date = source.get("publicationDate")
            if pub_date:
                year = pub_date.year if hasattr(pub_date, 'year') else str(pub_date)[:4]
            ws.cell(row_idx, 3, year or "")

            reason = source.get("exclusionReasoning") or source.get("inclusionReasoning") or ""
            ws.cell(row_idx, 4, reason)

            confidence = source.get("confidenceScore")
            if confidence is not None:
                ws.cell(row_idx, 5, f"{confidence:.0%}")

        # Format columns
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

    def _create_classification_matrix_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create Classification Matrix sheet with 0/1 indicators."""
        ws = wb.create_sheet("Classification Matrix", 4)

        facets = self._fetch_facets(study["id"])
        sources = self._fetch_included_sources(study["id"])

        # Build all possible categories
        all_categories = []
        category_to_facet = {}

        for facet in facets:
            categories = facet.get("categories", [])
            for cat in categories:
                all_categories.append(cat.get("name", ""))
                category_to_facet[cat.get("name", "")] = facet["id"]

        # Headers: ID, Title, then categories
        headers = ["ID", "Title"] + all_categories

        self._add_header_row(ws, headers)

        # Data
        for row_idx, source in enumerate(sources, 2):
            ws.cell(row_idx, 1, source.get("id", ""))
            ws.cell(row_idx, 2, source.get("title", ""))

            classifications = source.get("classifications", {})

            for col_idx, category in enumerate(all_categories, 3):
                facet_id = category_to_facet.get(category)

                if facet_id and facet_id in classifications:
                    facet_data = classifications[facet_id]
                    class_list = facet_data.get("classifications", [])

                    # Check if this category is in the classifications
                    has_category = any(c.get("categoryName") == category for c in class_list)
                    ws.cell(row_idx, col_idx, 1 if has_category else 0)
                else:
                    ws.cell(row_idx, col_idx, 0)

        # Add SUM row at bottom
        sum_row = len(sources) + 2
        ws.cell(sum_row, 1, "TOTAL")

        for col_idx in range(3, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(sum_row, col_idx, f"=SUM({col_letter}2:{col_letter}{len(sources) + 1})")

        # Format columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_data_dictionary_sheet(self, wb: Workbook, study: Dict[str, Any]) -> None:
        """Create Data Dictionary sheet."""
        ws = wb.create_sheet("Data Dictionary", 5)

        # Headers for data dictionary
        headers = ["Column Name", "Description", "Type", "Possible Values"]
        self._add_header_row(ws, headers)

        facets = self._fetch_facets(study["id"])

        row = 2

        # Standard columns from Included Sources
        standard_cols = [
            ("ID", "Unique identifier for the source", "Text", ""),
            ("Title", "Title of the source", "Text", ""),
            ("Authors", "Authors of the source (semicolon-separated)", "Text", ""),
            ("Year", "Publication year", "Number", ""),
            ("Venue", "Publication venue/conference/journal", "Text", ""),
            ("DOI", "Digital Object Identifier", "Text", ""),
        ]

        for name, desc, type_, values in standard_cols:
            ws.cell(row, 1, name)
            ws.cell(row, 2, desc)
            ws.cell(row, 3, type_)
            ws.cell(row, 4, values)
            row += 1

        # Facet columns
        for facet in facets:
            ws.cell(row, 1, facet.get("name", ""))
            ws.cell(row, 2, facet.get("description", "") or "")
            ws.cell(row, 3, f"Classification ({facet.get('type', '')})")

            categories = facet.get("categories", [])
            if categories:
                cat_names = [c.get("name", "") for c in categories]
                ws.cell(row, 4, "; ".join(cat_names))

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50

    def _add_header_row(self, ws, headers: List[str]) -> None:
        """Add a formatted header row."""
        fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

    def _add_row(self, ws, row: int, label: str, value: str, bold: bool = False) -> None:
        """Add a label-value row."""
        cell_label = ws.cell(row, 1, label)
        if bold:
            cell_label.font = Font(bold=True)

        cell_value = ws.cell(row, 2, value)
        if bold:
            cell_value.font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70
